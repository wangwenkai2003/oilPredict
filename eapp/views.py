

from sklearn.ensemble import AdaBoostRegressor, GradientBoostingRegressor
from sklearn.kernel_ridge import KernelRidge
from sklearn.preprocessing import LabelEncoder
from sklearn.svm import SVR

import os
import pandas as pd
import numpy as np
from django.http import JsonResponse, HttpResponse, FileResponse
from django.middleware.csrf import get_token
import openpyxl

from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split,GridSearchCV
from sklearn.linear_model import LinearRegression, SGDRegressor, Ridge, RidgeCV, LassoCV, BayesianRidge
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import PolynomialFeatures
from matplotlib import pyplot as plt
from sklearn.preprocessing import LabelEncoder

from matplotlib.pyplot import MultipleLocator
import base64
import joblib


def getcsrf(request):
     return JsonResponse({'csrftoken': get_token(request) or 'NOTPROVIDED'})

#处理文件
def upload_excel(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        # 保存上传的文件到服务器
        with open('uploaded_file.xlsx', 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        # 打开上传的文件并创建一个新的 Excel 文档
        wb = openpyxl.Workbook()
        ws = wb.active



        # 读取上传的文件内容并写入到新的 Excel 文档中
        with open('uploaded_file.xlsx', 'rb') as f:
            existing_wb = openpyxl.load_workbook(f, data_only=True)  # 使用 data_only 参数来确保读取数值而不是公式
            existing_ws = existing_wb.active
            for row in existing_ws.iter_rows():
                ws.append([cell.value for cell in row])

        # 保存新的 Excel 文档
        res =  handleExcel()
        os.remove('uploaded_file.xlsx')
        return JsonResponse(res)
    else:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

#对‘-’的数据进行清洗
def average_pressure(pressure_range):
    if not isinstance(pressure_range, str):
        # 如果不是字符串，则直接返回原始值（可能需要根据实际需求调整）
        return pressure_range
    if '-' in pressure_range:
        # 处理范围值
        range_parts = pressure_range.split('-')

        # 检查是否确实分割出了两个数值
        if len(range_parts) == 2:
            return (float(range_parts[0]) + float(range_parts[1])) / 2
        else:
            # 如果只有一个值，可以返回原值或者抛出异常，根据你的实际需求决定
            print(f"Warning: Invalid range value '{pressure_range}' - returning the original value.")
            return pressure_range
    elif '—' in pressure_range:
        # 处理范围值
        range_parts = pressure_range.split('—')

        # 检查是否确实分割出了两个数值
        if len(range_parts) == 2:
            return (float(range_parts[0]) + float(range_parts[1])) / 2
        else:
            # 如果只有一个值，可以返回原值或者抛出异常，根据你的实际需求决定
            print(f"Warning: Invalid range value '{pressure_range}' - returning the original value.")
            return pressure_range
    else:
        # 处理单个数值
        return pressure_range

# 定义一个函数来检查单元格是否包含时间格式数据
def is_time_anywhere(row):
    for value in row:
        str_value = str(value)
        try:
            if ':' in str_value :
                return True
        except ValueError:
            pass  # 继续检查下一个值

    return False

def is_number(x):
    # 检查是否为浮点数或整数
    return isinstance(x, (int, float))

def tmplate_excel(request):
    response = FileResponse(open('template.xlsx', 'rb'), as_attachment=True, filename='template.xlsx')
    return response

#模型训练
def handleExcel():
    data = pd.read_excel('uploaded_file.xlsx', sheet_name='Sheet2')

    le = LabelEncoder()

    # filcol = ['地层','井段','岩性','强度','可钻性级值','钻进工具','钻头型号','直径','钻进方式','钻压','转盘转速','转速','泵压','排量','扭矩','钻井液密度','井斜角','钻头压降','钻速']
    # 定义要保留的列名列表
    columns_to_keep = ['地层', '井段', '岩性', '强度MPa', '可钻性级值', '钻进工具', '钻头型号', '直径(mm)','钻进方式', '钻压/KN', '转盘转速/rpm', '螺杆转速/rpm', '泵压/Mpa', '排量(L/S)', '扭矩/(KN*m)', '钻井液密度g/cm3', '井斜角(°)', '钻头压降/Mpa', '钻速(m/h)']

    columns_to_average = ['强度MPa', '钻压/KN', '转盘转速/rpm', '螺杆转速/rpm', '泵压/Mpa',
                          '排量(L/S)', '钻井液密度g/cm3','扭矩/(KN*m)', '钻头压降/Mpa','井斜角(°)','钻速(m/h)']

    # 选择并保存需要的列
    data = data.loc[:, columns_to_keep]


    for column in columns_to_average:
        data[column] = data[column].apply(average_pressure)



    columns_to_encode = ['地层', '井段', '岩性', '钻头型号', '钻进工具', '钻进方式']


    for column in columns_to_encode:
        # 直接替换原始列
        data[column] = data[column].astype(str)
        data[column] = le.fit_transform(data[column])

    data = data[~data.apply(is_time_anywhere, axis=1)]

    data = data.iloc[1:]

    data = data.applymap(lambda x: x if is_number(x) else np.nan)
    data = data.dropna()


    # 设置Matplotlib字体
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定中文字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

    plt.xticks(fontsize=8)

    # 获取x轴
    data_h = pd.read_excel('uploaded_file.xlsx', sheet_name='Sheet2')
    column_headers = data_h.columns.tolist()
    # plt_x = ['地层', '井段', '岩性', '强度MPa', '可钻性级值', '钻进工具', '直径(mm)','钻进方式', '钻压/KN', '转盘转速/rpm', '螺杆转速/rpm', '泵压/Mpa', '排量(L/S)', '扭矩/(KN*m)', '钻井液密度g/cm3', '井斜角(°)', '钻头压降/Mpa']
    plt_x = ['地层', '井段', '岩性', '强度', '可钻性级值', '钻进工具', '直径','钻进方式', '钻压', '转盘转速', '螺杆转速', '泵压', '排量', '扭矩', '钻井液密度', '井斜角', '钻头压降']
    x_ticks = np.linspace(0, len(plt_x) - 1, len(plt_x))


    plt.xlim(-0.5, 17.5)

    # 创建一个每隔1个单位就有一个刻度的定位器
    x_major_locator = MultipleLocator(base=5)

    # 将此定位器应用到当前的x轴
    ax = plt.gca()
    # ax为两条坐标轴的实例
    ax.xaxis.set_major_locator(x_major_locator)

    # 设置 x 轴标签间距
    plt.xticks(x_ticks, plt_x, rotation=30)

    data.columns = data.columns.astype(str)
    columns_to_encode = [7,-1]
    columns_names = data.columns[columns_to_encode]
    encoded_col = pd.get_dummies(data[columns_names])
    selected_columns = data.drop(columns_names, axis=1)
    target = data.iloc[:, -1]
    x_train, x_test, y_train, y_test = train_test_split(selected_columns, target, random_state=62)
    l0 = train_linear_regression(x_train, y_train, x_test, y_test)
    l1 = train_ridge_regression(x_train, y_train, x_test, y_test)
    l2 = train_elasticnet_regression(x_train, y_train, x_test, y_test)
    l3 = train_lasson_regression(x_train, y_train, x_test, y_test)
    l4 = train_metrics(x_train, y_train, x_test, y_test)
    l = [l0, l1, l2, l3, l4]
    l_duibi = [l0[4],l1[4],l2[4],l3[4],l4[4]]
    x_min = l[0][2]
    l_min = l[0]
    best_index=-1;
    for index, value in enumerate(l):
        if value[2] < x_min:
            x_min = value[2]
            l_min = value
            best_index = index
    plt_y =  l_min[3]
    total_sum = np.sum(l_min[3])
    plt.bar(plt_x, plt_y)
    plt.subplots_adjust(bottom=0.2)


    plt.close()  # 关闭当前图形

    all_model_data = [l0[2], l1[2], l2[2], l3[2], l4[2]]
    duibi = ['线性回归', '岭回归', 'elstacinet回归', 'Lasso回归', '随机森林']
    colors = ['orange', 'orange', 'orange', 'orange','orange']
    plt.bar(range(len(duibi)), all_model_data, color=colors)
    plt.xticks(range(len(duibi)), duibi)
    for i in range(len(duibi)):
        plt.text(duibi[i], all_model_data[i] + 1, str(all_model_data[i]), ha='center')
    plt.close()  # 关闭当前图形


    # 读取图片文件并转换为 Base64 编码的字符串
    with open("metrics1.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    with open("metrics2.png", "rb") as image_file:
        all_model = base64.b64encode(image_file.read()).decode('utf-8')

    #所有模型的预测值与实际值
    par_pre=[[l0[0].tolist(),l0[1].tolist()],[l1[0].tolist(),l1[1].tolist()],[l2[0].tolist(),l2[1].tolist()],[l3[0].tolist(),l3[1].tolist()],[l4[0].tolist(),l4[1].tolist()]]


    # 将l_min中的每个元素转换为列表
    l_min[0]=l_min[0].tolist()
    l_min[1]=l_min[1].tolist()
    l_min[3]=l_min[3].tolist()
    # 创建一个字典，包含所有模型名称、数据、模型、最佳模型、最小值、预训练参数和最佳模型名称
    result = {
        'all_model_name':duibi,
        'all_model_data':all_model_data,
        'all_model':all_model,
        'best_model': image_data,
        'l_min': l_min,
        'pre_par': par_pre,
        'best_name':duibi[best_index],
        'l_duibi':l_duibi
    }
    return result


def filePredict(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        model_index = request.POST.get('modelIndex', '0')

        le = LabelEncoder()
        # 保存文件
        with open('predict.xlsx', 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        # 加载模型
        select_model = f"model{model_index}.pkl"
        estimator = joblib.load(select_model)

        # 指定需要保留的列名
        columns_to_keep = ['地层', '井段', '岩性', '强度MPa', '可钻性级值', '钻进工具', '钻头型号', '直径(mm)',
                           '钻进方式', '钻压/KN', '转盘转速/rpm', '螺杆转速/rpm', '泵压/Mpa', '排量(L/S)',
                           '扭矩/(KN*m)', '井斜角(°)', '钻头压降/Mpa']

        # 读取Excel文件，假设第一行为列名
        data = pd.read_excel('predict.xlsx', sheet_name='Sheet2')

        # 筛选指定列
        data = data[columns_to_keep]

        # 数据预处理
        columns_to_average = ['强度MPa', '钻压/KN', '转盘转速/rpm', '螺杆转速/rpm', '泵压/Mpa',
                              '排量(L/S)',  '扭矩/(KN*m)', '钻头压降/Mpa', '井斜角(°)']

        for column in columns_to_average:
            data[column] = data[column].apply(average_pressure)

        columns_to_encode = ['地层', '井段', '岩性', '钻头型号', '钻进工具', '钻进方式']

        for column in columns_to_encode:
            data[column] = data[column].astype(str)
            data[column] = le.fit_transform(data[column])

        data = data.applymap(lambda x: x if is_number(x) else np.nan)
        data = data.dropna()

        # 过滤特定条件
        data = data[~data.apply(is_time_anywhere, axis=1)]

        # 重置索引并且去除所有列名，为模型预测准备
        filtered_indices = data.index
        data.reset_index(drop=True, inplace=True)
        data.columns = range(data.shape[1])

        # 模型预测
        y_predict = estimator.predict(data)

        # 读取原始数据集用于添加预测结果
        original_data = pd.read_excel('predict.xlsx', sheet_name='Sheet2')

        # 创建一个全为NaN的序列，长度与原始数据集的行数相同
        prediction_series = pd.Series([np.nan] * len(original_data), index=original_data.index)

        # 更新原始数据集中的预测结果
        prediction_series.loc[filtered_indices] = y_predict
        original_data['预测钻速(m/h)'] = prediction_series

        # 保存更新后的数据集
        updated_file_path = 'updated_predict.xlsx'
        original_data.to_excel(updated_file_path, index=False)

        # 返回文件
        response = FileResponse(open(updated_file_path, 'rb'), as_attachment=True, filename='updated_predict.xlsx')

        # 注意：这里我们不立即删除文件，以避免在文件还在传输过程中被删除的问题
        # 可能需要在文件成功下载后，通过其他机制来删除这些文件

        return response

# 线性回归
def train_linear_regression(x_train, y_train, x_test, y_test):
    # 训练线性回归模型
    estimator = LinearRegression()
    estimator.fit(x_train, y_train)
    # 6）模型评估
    # score = estimator.score(x_test, y_test)
    # print("线性回归结果:")
    # print("正确率为：", score)
    y_predict = np.round(estimator.predict(x_test),2)
    print("线性回归结果:")
    print("预测值：\n", y_predict)
    print("真实值：\n", np.array(y_test))
    error = mean_squared_error(y_test, y_predict)
    print("正规方程-均方误差为：\n", error)
    plt.figure()
    plt.scatter(y_test, y_predict)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
    plt.xlabel('真实值')
    plt.ylabel('预测值')
    plt.savefig('lineImg')
    plt.close()
    with open("lineImg.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    joblib.dump(estimator, "model0.pkl")
    l = [y_predict, np.round(np.array(y_test),2), error, estimator.coef_,image_data]
    return l

#岭回归
def train_ridge_regression(x_train, y_train, x_test,y_test):
    # 特征标准化
    transfer = StandardScaler()
    x_train = transfer.fit_transform(x_train)
    x_test = transfer.transform(x_test)

    alphas_to_test = np.linspace(0.001,1)
    # 训练线性回归模型
    estimator = RidgeCV(alphas=alphas_to_test,store_cv_values=True)
    estimator.fit(x_train, y_train)

    best_alpha = estimator.alpha_
    # print(x_train)

    # 使用最佳参数重新训练模型
    best_estimator = RidgeCV(alphas=[best_alpha], store_cv_values=True)
    best_estimator.fit(x_train, y_train)

   #岭回归
def train_ridge_regression(x_train, y_train, x_test,y_test):
    # 特征标准化
    transfer = StandardScaler()
    x_train = transfer.fit_transform(x_train)
    x_test = transfer.transform(x_test)

    alphas_to_test = np.linspace(0.001,1)
    # 训练线性回归模型
    estimator = RidgeCV(alphas=alphas_to_test,store_cv_values=True)
    estimator.fit(x_train, y_train)

    best_alpha = estimator.alpha_
    # print(x_train)

    # 使用最佳参数重新训练模型
    best_estimator = RidgeCV(alphas=[best_alpha], store_cv_values=True)
    best_estimator.fit(x_train, y_train)

#岭回归
def train_ridge_regression(x_train, y_train, x_test,y_test):
    # 特征标准化
    transfer = StandardScaler()
    x_train = transfer.fit_transform(x_train)
    x_test = transfer.transform(x_test)

    alphas_to_test = np.linspace(0.001,1)
    # 训练线性回归模型
    estimator = RidgeCV(alphas=alphas_to_test,store_cv_values=True)
    estimator.fit(x_train, y_train)

    best_alpha = estimator.alpha_
    # print(x_train)

    # 使用最佳参数重新训练模型
    best_estimator = RidgeCV(alphas=[best_alpha], store_cv_values=True)
    best_estimator.fit(x_train, y_train)

    # 保存模型
    joblib.dump(best_estimator, "model1.pkl")
    # 计算测试集的拟合分数
    score = best_estimator.score(x_test, y_test)
    # 打印岭回归的结果
    print("岭回归结果:")
    y_predict = np.round(best_estimator.predict(x_test), 2)
    print("预测值：\n", y_predict)
    print("实际值：\n", np.floor(np.round(np.array(y_test), 2) * 100) / 100)
    # 计算均方误差
    error = mean_squared_error(y_test, y_predict)
    # 绘制散点图
    plt.figure()
    plt.scatter(y_test, y_predict)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
    plt.xlabel('真实值')
    plt.ylabel('预测值')
    # 保存图片
    plt.savefig('RidgeImg')
    plt.close()
    # 读取图片并编码为base64
    with open("RidgeImg.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    print("正规方程-均方误差为：\n", error)
    # 返回预测值，实际值，均方误差，系数，图片数据
    l = [y_predict, np.round(np.array(y_test), 2), error, estimator.coef_, image_data]
    return l

#多项式回归
def train_polynomial_regression(x_train, y_train, x_test,y_test):
    # 5. 使用最优的多项式阶数进行模型训练和预测
    # best_degree = grid_search.best_params_['polynomial_features__degree']
    # 创建pipeline，包括多项式特征转换和线性回归模型
    # bemodel = make_pipeline(PolynomialFeatures(), LinearRegression())
    # #
    # # 设置参数网格
    # param_grid = {'polynomialfeatures__degree': np.arange(1, 10)}
    #
    # # 使用GridSearchCV进行交叉验证
    # grid = GridSearchCV(bemodel, param_grid, cv=5)
    # grid.fit(x_train, y_train)

    poly_features = PolynomialFeatures(degree=5)
    X_train_poly = poly_features.fit_transform(x_train)
    X_test_poly = poly_features.transform(x_test)

    model = LinearRegression()
    model.fit(X_train_poly, y_train)

    # y_pred = model.predict(X_test_poly)
    # print(y_pred)
    joblib.dump(model, "model2.pkl")
    # 6）模型评估
    score = model.score(X_test_poly, y_test)
    print("多项式回归结果:")
    y_predict = np.round(model.predict(X_test_poly), 2)
    print("预测值：\n", y_predict)
    print("实际值：\n", np.array(y_test))
    plt.figure()
    plt.scatter(y_test, y_predict)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
    plt.xlabel('真实值')
    plt.ylabel('预测值')
    plt.savefig('PolyImg')
    plt.close()
    with open("PolyImg.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    error = mean_squared_error(y_test, y_predict)
    l = [y_predict, np.round(np.array(y_test),2), error, model.coef_,image_data]
    return l

#使用lasso回归
def train_lasson_regression(x_train, y_train, x_test,y_test):
    # lasso回归
    lasso_clf = LassoCV()
    lasso_clf.fit(x_train, y_train)
    joblib.dump(lasso_clf, "model3.pkl")
    # 预测
    y_predict = np.round(lasso_clf.predict(x_test), 2)
    # 打印预测结果
    print("lasson回归结果:")
    print("预测值：\n", y_predict)
    print("实际值：\n", np.array(y_test))
    # 绘制散点图和直线
    plt.figure()
    plt.scatter(y_test, y_predict)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
    # 设置x轴和y轴的标签
    plt.xlabel('真实值')
    plt.ylabel('预测值')
    # 保存图片
    plt.savefig('lassoImg')
    # 关闭图形
    plt.close()

    # 读取图片并编码
    with open("lassoImg.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    # 计算均方误差
    error = mean_squared_error(y_test, y_predict)
    print("正规方程-均方误差为：\n", error)
    print(lasso_clf.score(x_test,y_test))
    # 返回结果
    l = [y_predict, np.round(np.array(y_test),2), error, lasso_clf.coef_,image_data]
    return l

#使用随机森林
def train_metrics(x_train, y_train, x_test,y_test):
    '''
    :param x_train: 训练集特征
    :param y_train: 训练集标签
    :param x_test: 测试集特征
    :param y_test: 测试集标签
    :return: 预测结果，真实结果，均方误差，特征重要性，混淆矩阵图
    '''
    from sklearn.ensemble import RandomForestRegressor
    from sklearn import metrics
    # 初始化随机森林回归器
    rf_regressor = RandomForestRegressor(n_estimators=10, random_state=25)
    # 训练随机森林回归器
    rf_regressor.fit(x_train, y_train)
    # 将模型保存为pkl文件
    joblib.dump(rf_regressor, "model4.pkl")
    # 预测
    y_pred = rf_regressor.predict(x_test)
    # 计算均方误差
    mse = metrics.mean_squared_error(y_test, y_pred)
    # 获取预测结果，保留两位小数
    predicted_target = np.round(rf_regressor.predict(x_test),2)
    # 绘制散点图
    plt.figure()
    plt.scatter(y_test, y_pred)
    # 绘制连接线
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
    # 设置轴标签
    plt.xlabel('真实值')
    plt.ylabel('预测值')
    # 保存混淆矩阵图
    plt.savefig('随机森林')
    plt.close()
    # 将混淆矩阵图转换为base64编码
    with open("随机森林.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')

    # 返回预测结果，真实结果，均方误差，特征重要性，混淆矩阵图
    return [predicted_target, np.round(np.array(y_test),2), mse, np.array(rf_regressor.feature_importances_),image_data]

#使用elasticnet回归
def train_elasticnet_regression(x_train, y_train, x_test,y_test):
    '''
    此函数使用elasticnet回归训练模型，并将结果存储在'model2.pkl'文件中。
    同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
    '''
    from sklearn.linear_model import ElasticNet
    #创建elasticnet回归模型
    elastic_net = ElasticNet(random_state=0)
    #使用训练数据拟合模型
    elastic_net.fit(x_train, y_train)
    #将模型存储在'model2.pkl'文件中
    joblib.dump(elastic_net, "model2.pkl")
    #使用测试数据进行预测
    y_predict = np.round(elastic_net.predict(x_test),2)
    #计算均方误差
    error = mean_squared_error(y_test, y_predict)
    print("elasticnet回归结果:")
    print("预测值：\n", y_predict)
    print("实际值：\n", np.array(y_test))
    print("正规方程-均方误差为：\n", error)
    #绘制预测值和实际值的散点图
    plt.figure()
    plt.scatter(y_test, y_predict)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
    plt.xlabel('真实值')
    plt.ylabel('预测值')
    plt.savefig('elastic')
    plt.close()
    #将散点图存储在'elastic.png'文件中
    with open("elastic.png", "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    #返回预测值、实际值、均方误差和模型系数，以及图像数据
    l = [y_predict, np.round(np.array(y_test), 2), error, np.array(elastic_net.coef_), image_data]
    return l
#支持向量回归（SVR）
# def train_svr_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用支持向量回归（SVR）训练模型，并将结果存储在'model_svr.pkl'文件中。
#     同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
#     """
#     svr = SVR(kernel='linear', C=1.0, epsilon=0.1)
#     svr.fit(x_train, y_train)
#     joblib.dump(svr, "model_svr.pkl")
#
#     y_predict = np.round(svr.predict(x_test), 2)
#     error = mean_squared_error(y_test, y_predict)
#
#     print("支持向量回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.array(y_test))
#     print("均方误差为：\n", error)
#
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.savefig('svr')
#     plt.close()
#
#     with open("svr.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     l = [y_predict, np.round(np.array(y_test), 2), error, np.array(svr.coef_), image_data]
#     return l
# # Adaptive Boosting回归
# def train_adaboost_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用Adaptive Boosting回归（AdaBoostRegressor）训练模型，并将结果存储在'model_adaboost.pkl'文件中。
#     同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
#     """
#     ada_regressor = AdaBoostRegressor(random_state=0)
#     ada_regressor.fit(x_train, y_train)
#     joblib.dump(ada_regressor, "model_adaboost.pkl")
#
#     y_predict = np.round(ada_regressor.predict(x_test), 2)
#     error = mean_squared_error(y_test, y_predict)
#
#     print("Adaptive Boosting回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.array(y_test))
#     print("均方误差为：\n", error)
#
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.savefig('adaboost')
#     plt.close()
#
#     with open("adaboost.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     l = [y_predict, np.round(np.array(y_test), 2), error, np.array(ada_regressor.feature_importances_), image_data]
#     return l
# #梯度提升回归树
# def train_gradient_boosting_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用梯度提升回归树（GradientBoostingRegressor）训练模型，并将结果存储在'model_gradient_boosting.pkl'文件中。
#     同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
#     """
#     gb_regressor = GradientBoostingRegressor(random_state=0)
#     gb_regressor.fit(x_train, y_train)
#     joblib.dump(gb_regressor, "model_gradient_boosting.pkl")
#
#     y_predict = np.round(gb_regressor.predict(x_test), 2)
#     error = mean_squared_error(y_test, y_predict)
#
#     print("梯度提升回归树结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.array(y_test))
#     print("均方误差为：\n", error)
#
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.savefig('gradient_boosting')
#     plt.close()
#
#     with open("gradient_boosting.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     l = [y_predict, np.round(np.array(y_test), 2), error, np.array(gb_regressor.feature_importances_), image_data]
#     return l
#
#
# def train_gradient_boosting_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用梯度提升回归（Gradient Boosting Regression）训练模型，并将结果存储在'model_gbr.pkl'文件中。
#     同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
#     """
#     gbr = GradientBoostingRegressor(random_state=42)
#     gbr.fit(x_train, y_train)
#     joblib.dump(gbr, "model_gbr.pkl")
#
#     y_predict = np.round(gbr.predict(x_test), 2)
#     error = mean_squared_error(y_test, y_predict)
#
#     print("梯度提升回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.array(y_test))
#     print("均方误差为：\n", error)
#
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.savefig('gradient_boosting')
#     plt.close()
#
#     with open("gradient_boosting.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     l = [y_predict, np.round(np.array(y_test), 2), error, None, image_data]
#     return l
#
#
# def train_support_vector_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用支持向量回归（Support Vector Regression）训练模型，并将结果存储在'model_svr.pkl'文件中。
#     同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
#     """
#     svr = SVR(kernel='linear', C=1.0, epsilon=0.1)
#     svr.fit(x_train, y_train)
#     joblib.dump(svr, "model_svr.pkl")
#
#     y_predict = np.round(svr.predict(x_test), 2)
#     error = mean_squared_error(y_test, y_predict)
#
#     print("支持向量回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.array(y_test))
#     print("均方误差为：\n", error)
#
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.savefig('support_vector')
#     plt.close()
#
#     with open("support_vector.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     l = [y_predict, np.round(np.array(y_test), 2), error, None, image_data]
#     return l
#
# def train_neural_network_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用神经网络回归（Neural Network Regression）训练模型，并将结果存储在'model_nn.pkl'文件中。
#     同时，它还计算了模型的均方误差（MSE）并绘制了预测值和实际值的散点图。
#     """
#     nn_reg = MLPRegressor(hidden_layer_sizes=(30,), activation='relu', solver='adam', random_state=42)
#     nn_reg.fit(x_train, y_train)
#     joblib.dump(nn_reg, "model_nn.pkl")
#
#     y_predict = np.round(nn_reg.predict(x_test), 2)
#     error = mean_squared_error(y_test, y_predict)
#
#     print("神经网络回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.array(y_test))
#     print("均方误差为：\n", error)
#
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k')
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.savefig('neural_network')
#     plt.close()
#
#     with open("neural_network.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     l = [y_predict, np.round(np.array(y_test), 2), error, None, image_data]
#     return l
#
#
#
#
# import numpy as np
# from sklearn.tree import DecisionTreeRegressor
# from sklearn.metrics import mean_squared_error
# import matplotlib.pyplot as plt
# import base64
#
# def train_decision_tree_regression(x_train, y_train, x_test, y_test):
#     """
#     此函数使用决策树回归模型（Decision Tree Regression）对给定的训练数据进行训练，并将训练得到的模型保存至 'model_dt.pkl' 文件。同时，函数计算模型在测试集上的预测性能指标——均方误差（MSE），并绘制预测值与实际值的散点图，以直观展示模型预测效果。
#
#     参数:
#     - x_train (numpy.ndarray): 训练集特征数据，二维数组形式。
#     - y_train (numpy.ndarray): 训练集目标变量数据，一维数组形式。
#     - x_test (numpy.ndarray): 测试集特征数据，二维数组形式，与训练集特征维度相同。
#     - y_test (numpy.ndarray): 测试集目标变量数据，一维数组形式，与训练集目标变量长度相同。
#
#     返回:
#     - list: 包含以下元素的列表：
#         1. 预测值数组（numpy.ndarray），保留两位小数；
#         2. 实际值数组（numpy.ndarray），同样保留两位小数；
#         3. 均方误差（float），表示预测值与实际值之差平方的平均值；
#         4. None（placeholder）；
#         5. 散点图的Base64编码字符串，用于在网络环境中直接展示图形。
#     """
#
#     # 初始化决策树回归模型，设置随机种子确保可复现性
#     dt_reg = DecisionTreeRegressor(random_state=42)
#
#     # 使用训练数据拟合决策树回归模型
#     dt_reg.fit(x_train, y_train)
#
#     # 将训练好的模型持久化保存到磁盘
#     joblib.dump(dt_reg, "model_dt.pkl")
#
#     # 对测试集数据进行预测，结果保留两位小数
#     y_predict = np.round(dt_reg.predict(x_test), 2)
#
#     # 计算预测值与实际值之间的均方误差
#     error = mean_squared_error(y_test, y_predict)
#
#     # 输出模型预测结果及均方误差
#     print("决策树回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.round(np.array(y_test), 2))
#     print("均方误差（MSE）：\n", error)
#
#     # 绘制预测值与实际值的散点图，并添加最佳拟合线（对角线）
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], '--k', label='Perfect Fit')  # 最佳拟合线
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.legend()
#     plt.title('决策树回归：预测值与实际值散点图')
#     plt.savefig('decision_tree.png')  # 保存图表至本地文件
#
#     # 关闭当前绘图窗口
#     plt.close()
#
#     # 将散点图转换为Base64编码以便在网络环境中直接展示
#     with open("decision_tree.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     # 构建并返回包含预测值、实际值、均方误差及散点图Base64编码的列表
#     return [y_predict, np.round(np.array(y_test), 2), error, None, image_data]
#
# def train_gam_regression(x_train, y_train, x_test, y_test):
#     # 初始化GAM模型，指定光滑函数类型（例如，s(0)代表一个未正则化的B样条基）
#     gam_reg = GAM(random_state=42)
#
#     # 使用训练数据拟合GAM模型
#     gam_reg.fit(x_train, y_train)
#
#     # 将训练好的模型持久化保存到磁盘
#     joblib.dump(gam_reg, "model_gam.pkl")
#
#     # 对测试集数据进行预测，结果保留两位小数
#     y_predict = np.round(gam_reg.predict(x_test), 2)
#
#     # 计算预测值与实际值之间的均方误差
#     error = mean_squared_error(y_test, y_predict)
#
#     # 输出模型预测结果及均方误差
#     print("GAM回归结果:")
#     print("预测值：\n", y_predict)
#     print("实际值：\n", np.round(np.array(y_test), 2))
#     print("均方误差（MSE）：\n", error)
#
#     # 绘制预测值与实际值的散点图
#     plt.figure()
#     plt.scatter(y_test, y_predict)
#     plt.xlabel('真实值')
#     plt.ylabel('预测值')
#     plt.title('GAM回归：预测值与实际值散点图')
#     plt.savefig('gam_regression.png')  # 保存图表至本地文件
#
#     # 关闭当前绘图窗口
#     plt.close()
#
#     # 将散点图转换为Base64编码以便在网络环境中直接展示
#     with open("gam_regression.png", "rb") as image_file:
#         image_data = base64.b64encode(image_file.read()).decode('utf-8')
#
#     # 构建并返回包含预测值、实际值、均方误差及散点图Base64编码的列表
#     return [y_predict, np.round(np.array(y_test), 2), error, None, image_data]
#
# from sklearn.preprocessing import PolynomialFeatures
# from sklearn.linear_model import LinearRegression
#
#
#
#
# def train_bys_regression(x_train, y_train, x_test, y_test):
#     # 定义贝叶斯回归模型
#     bayes_regression = BayesianRidge()
#
#     # 使用训练集拟合模型
#     bayes_regression.fit(x_train, y_train)
#
#     # 保存模型到文件
#     joblib.dump(bayes_regression, 'bayes_regression_model.pkl')
#
#     # 对测试集数据进行预测
#     y_pred = bayes_regression.predict(x_test)
#
#     # 计算MSE等评价指标
#     mse = mean_squared_error(y_test, y_pred)
#     print(f"Mean Squared Error (MSE): {mse}")
#
#     # 可视化结果
#     plt.figure(figsize=(8, 6))
#     plt.scatter(x_test, y_test, label="Actual values")
#     plt.plot(x_test, y_pred, color="red", label="Predicted values")
#     plt.xlabel("Test data")
#     plt.ylabel("Target variable")
#     plt.title("Bayesian Regression: Actual vs Predicted")
#     plt.legend()
#     plt.savefig("bayes_regression_plot.png")
#     plt.show()
#
# def train_AdaBoost_regression(x_train, y_train, x_test, y_test):
#     # 定义AdaBoost.R2模型，这里使用默认的决策树作为基学习器
#     ada_boost = AdaBoostRegressor(base_estimator=None, n_estimators=50)
#
#     # 使用训练集拟合模型
#     ada_boost.fit(x_train, y_train)
#
#     # 保存模型到文件
#     joblib.dump(ada_boost, 'ada_boost_regression_model.pkl')
#
#     # 对测试集数据进行预测
#     y_pred = ada_boost.predict(x_test)
#
#     # 计算MSE等评价指标
#     mse = mean_squared_error(y_test, y_pred)
#     print(f"Mean Squared Error (MSE): {mse}")
#
#     # 可视化结果
#     plt.figure(figsize=(8, 6))
#     plt.scatter(x_test, y_test, label="Actual values")
#     plt.plot(x_test, y_pred, color="red", label="Predicted values")
#     plt.xlabel("Test data")
#     plt.ylabel("Target variable")
#     plt.title("AdaBoost.R2 Regression: Actual vs Predicted")
#     plt.legend()
#     plt.savefig("ada_boost_regression_plot.png")
#     plt.show()
#
# def train_kernel_regression(x_train, y_train, x_test, y_test):
#     # 定义核岭回归模型，这里使用RBF核
#     kernel = 'rbf'
#     alpha = 1.0
#     gamma = 'scale'  # 或者指定一个固定的值，如 gamma=0.1
#     kernel_ridge = KernelRidge(alpha=alpha, kernel=kernel, gamma=gamma)
#
#     # 使用训练集拟合模型
#     kernel_ridge.fit(x_train, y_train)
#
#     # 保存模型到文件
#     joblib.dump(kernel_ridge, 'kernel_ridge_model.pkl')
#
#     # 对测试集数据进行预测
#     y_pred = kernel_ridge.predict(x_test)
#
#     # 计算MSE等评价指标
#     mse = mean_squared_error(y_test, y_pred)
#     print(f"Mean Squared Error (MSE): {mse}")
#
#     # 可视化结果
#     plt.figure(figsize=(8, 6))
#     plt.scatter(x_test, y_test, label="Actual values")
#     plt.plot(x_test, y_pred, color="red", label="Predicted values")
#     plt.xlabel("Test data")
#     plt.ylabel("Target variable")
#     plt.title("Kernel Ridge Regression: Actual vs Predicted")
#     plt.legend()
#     plt.savefig("kernel_ridge_regression_plot.png")
#     plt.show()